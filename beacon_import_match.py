"""
Cross-reference BeaconImportData.xlsx (July 2025 commercial snapshot)
against HELM's current `clients` table. Reports how many accounts match,
how many are missing, and breaks down by company.
"""
import json, urllib.parse, urllib.request
from openpyxl import load_workbook

SUPABASE_URL  = 'https://iiqlhqtyyqctgupalrlc.supabase.co'
SUPABASE_ANON = 'sb_publishable_JAASDdg8OERvpDTnEeLd5Q_Re48qR1I'
IN_PATH = r"C:\Users\theco\Downloads\BeaconImportData.xlsx"


def fetch_clients_by_ids(ids):
    if not ids: return {}
    # PostgREST in.() has a URL-length ceiling — batch in groups of 200.
    out = {}
    for i in range(0, len(ids), 200):
        chunk = ids[i:i+200]
        qs = urllib.parse.urlencode({
            "select": "client_id,client_name,address,phone,email,status",
            "client_id": "in.({})".format(",".join(chunk)),
        })
        url = f"{SUPABASE_URL}/rest/v1/clients?{qs}"
        req = urllib.request.Request(url, headers={
            "apikey": SUPABASE_ANON,
            "Authorization": f"Bearer {SUPABASE_ANON}",
        })
        with urllib.request.urlopen(req, timeout=30) as resp:
            rows = json.loads(resp.read().decode("utf-8"))
        for r in rows: out[r["client_id"]] = r
    return out


def collect_sheet_accounts(ws, acct_col=1, name_col=5, addr_col=6):
    rows = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, acct_col).value
        if v is None: continue
        if isinstance(v, (int, float)):
            acct = str(int(v))
        else:
            s = str(v).strip()
            if not s.isdigit(): continue
            acct = s
        rows.append({
            "acct": acct,
            "name": ws.cell(r, name_col).value or "",
            "addr": ws.cell(r, addr_col).value or "",
        })
    return rows


def main():
    wb = load_workbook(IN_PATH, data_only=True)
    reis   = collect_sheet_accounts(wb["Reis"])
    santos = collect_sheet_accounts(wb["Santos"])
    print(f"Reis sheet:   {len(reis)} accounts")
    print(f"Santos sheet: {len(santos)} accounts")
    print(f"Total:        {len(reis)+len(santos)}\n")

    all_ids = list({r["acct"] for r in reis} | {r["acct"] for r in santos})
    print(f"Unique acct #s: {len(all_ids)}")
    print(f"Looking up in HELM…\n")
    matched = fetch_clients_by_ids(all_ids)
    print(f"  matched: {len(matched)} / {len(all_ids)}")
    print(f"  missing: {len(all_ids) - len(matched)}\n")

    def report(label, lst):
        m = sum(1 for r in lst if r["acct"] in matched)
        u = len(lst) - m
        print(f"  {label:8s}  matched {m:>3} / {len(lst):>3}   ({100*m/max(1,len(lst)):5.1f}%)   missing {u}")
        return m, u

    print("By company:")
    report("Reis",   reis)
    report("Santos", santos)

    # Status breakdown of matched
    print("\nStatus of matched HELM accounts:")
    status_counts = {}
    for v in matched.values():
        s = v.get("status") or "(none)"
        status_counts[s] = status_counts.get(s, 0) + 1
    for s, c in sorted(status_counts.items(), key=lambda kv: -kv[1]):
        print(f"  {s:>10s}: {c}")

    # Missing accounts — first 20 of each
    print("\nMissing from HELM — Reis (first 20):")
    miss_reis = [r for r in reis if r["acct"] not in matched]
    for r in miss_reis[:20]:
        print(f"  #{r['acct']}  {r['name'][:40]:40s}  {r['addr']}")
    if len(miss_reis) > 20:
        print(f"  …and {len(miss_reis)-20} more")

    print("\nMissing from HELM — Santos (first 20):")
    miss_santos = [r for r in santos if r["acct"] not in matched]
    for r in miss_santos[:20]:
        print(f"  #{r['acct']}  {r['name'][:40]:40s}  {r['addr']}")
    if len(miss_santos) > 20:
        print(f"  …and {len(miss_santos)-20} more")


if __name__ == "__main__":
    main()
