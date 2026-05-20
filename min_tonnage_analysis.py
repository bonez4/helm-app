"""
Walk-in (C&D) minimum-tonnage analysis for calendar 2025.

Reads the SpreadsheetML scale export and emits an Excel workbook where the
user can toggle a minimum-tonnage threshold (and optional rate override)
and see live revenue uplift across all walk-in tickets.

Definitions (matched to HELM's parser, index.html ~line 6202):
- A "walk-in" ticket is any ticket that ISN'T:
    * an intercompany rolloff (ZZISLAND / ZZEASTEND)
    * a pile pickup (ZZDELTA) or outbound (ZZREIS)
    * a non-ZZ ticket that carries a rolloff service line
      (ROLLOFF DELIVERY / EMPTY ROLLOFF / ROLLOFF DOUBLE DROP /
       MOVE ROLLOFF ON JOB SITE)
  Walk-ins include both ZZ-cash codes (ZZTNT etc.) and contractor
  drive-throughs billed to a real account.

Output: Tonnage_Analysis_2025.xlsx with three sheets:
  - Dashboard  : toggle cells + headline KPIs + tons-bucket distribution
  - Tickets    : every walk-in ticket in 2025, with live formulas
  - Monthly    : month rollup of current vs adjusted revenue
"""
import re
import sys
from collections import defaultdict
from datetime import date
from html import unescape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

XLS_PATH = r"C:\Users\theco\OneDrive\Desktop\Nantucket\helm-app\1776188206521-JAN12025TOAPR132026.xls"
OUT_PATH = r"C:\Users\theco\OneDrive\Desktop\Nantucket\helm-app\Tonnage_Analysis_2025.xlsx"

ROLLOFF_SVC_KWS = (
    "ROLLOFF DELIVERY",
    "ROLLOFF DOUBLE DROP",
    "EMPTY ROLLOFF",
    "EMPTY ROLL-OFF",
    "MOVE ROLLOFF ON JOB SITE",
)
IC_CODES = {"ZZISLAND", "ZZEASTEND", "ZZDELTA", "ZZREIS"}

# Per-unit disposal fees, mirrored from HELM
DISPOSAL_FEES = {
    "MATTRESS LANDFILL DISPOSAL FEE": 50.00,
    "STOVE DISHWASH APPLIANCE DISP":  11.00,
    "FRION LANDFILL DISPOSAL FEE":    45.00,
    "MONITOR LANDFILL DISPOSAL FEE":  16.50,
    "DUMPTIRE":                       16.50,
}


def parse_rows(xml_text):
    """Parse SpreadsheetML rows into list-of-lists, preserving cell index gaps."""
    xml_text = re.sub(r"<\?[^?]*\?>", "", xml_text)
    row_rx = re.compile(r"<(?:ss:)?Row[^>]*>([\s\S]*?)</(?:ss:)?Row>", re.IGNORECASE)
    cell_rx = re.compile(r"<(?:ss:)?Cell([^>]*)>([\s\S]*?)</(?:ss:)?Cell>", re.IGNORECASE)
    data_rx = re.compile(r"<(?:ss:)?Data[^>]*>([\s\S]*?)</(?:ss:)?Data>", re.IGNORECASE)
    idx_rx = re.compile(r'ss:Index="(\d+)"')

    rows = []
    for m in row_rx.finditer(xml_text):
        rc = m.group(1)
        vals, ci = [], 0
        for cm in cell_rx.finditer(rc):
            idxm = idx_rx.search(cm.group(1))
            if idxm:
                t = int(idxm.group(1)) - 1
                while ci < t:
                    vals.append("")
                    ci += 1
            dm = data_rx.search(cm.group(2))
            vals.append(unescape(dm.group(1).strip()) if dm else "")
            ci += 1
        rows.append(vals)
    return rows


def mdy_to_iso(mdy):
    parts = mdy.split("/")
    if len(parts) != 3:
        return None
    return f"{parts[2]}-{int(parts[0]):02d}-{int(parts[1]):02d}"


def extract_tickets(rows):
    """Walk the rows and build a per-ticket dict, mirroring HELM's logic."""
    tickets = {}
    code_stack = []
    current_code = None
    parent_code = None
    current_name = None  # display name from the customer header
    code_to_name = {}

    def is_header(first):
        if not first or " - " not in first:
            return False
        bad = ("Totals", "Tickets:", "TIPFEES", "ROLLOFFS", "No Order",
               "Date Range", "Report")
        if any(b in first for b in bad):
            return False
        if first.startswith("Page "):
            return False
        if first.strip() == "Date":
            return False
        return True

    def is_totals(first):
        return first and " - " in first and "Totals" in first \
            and "TIPFEES" not in first and "Non Order" not in first

    for vals in rows:
        non_empty = [v for v in vals if v and v.strip()]
        if not non_empty:
            continue
        first = non_empty[0]

        if is_header(first):
            parts = first.split(" - ", 1)
            code = parts[0].strip()
            name = parts[1].strip() if len(parts) > 1 else ""
            code_stack.append(code)
            current_code = code
            parent_code = code_stack[0] if len(code_stack) > 1 else None
            if name:
                code_to_name[code] = name
            current_name = code_to_name.get(parent_code or current_code, name)

        if is_totals(first):
            code = first.split(" - ")[0].strip()
            if code in code_stack:
                idx = len(code_stack) - 1 - code_stack[::-1].index(code)
                code_stack.pop(idx)
            current_code = code_stack[-1] if code_stack else None
            parent_code = code_stack[0] if len(code_stack) > 1 else None
            current_name = code_to_name.get(parent_code or current_code or "", "")

        effective_code = parent_code if (parent_code and parent_code.startswith("ZZ")) else current_code
        if not effective_code:
            continue

        flat_upper = " ".join(vals).upper()
        data_vals = [v for v in vals if v not in (None, "", )]

        # REIS SITE TRANSFER FEE → tonnage line
        if "REIS SITE TRANSFER FEE" in flat_upper:
            fi = next((i for i, v in enumerate(data_vals)
                       if v and "REIS SITE TRANSFER FEE" in v), -1)
            if fi != -1 and fi + 2 < len(data_vals):
                ticket = data_vals[1] or "?"
                try:
                    net = float(data_vals[fi + 2])
                except (TypeError, ValueError):
                    net = 0.0
                rate = 0.0
                if fi + 3 < len(data_vals):
                    rm = re.search(r"\$([0-9.]+)", data_vals[fi + 3] or "")
                    if rm:
                        rate = float(rm.group(1))
                # date is in data_vals[0] typically
                d = None
                for v in data_vals:
                    if v and re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", v):
                        d = v
                        break

                t = tickets.setdefault(ticket, {
                    "tons": 0.0, "rate": 0.0, "service": None,
                    "code": effective_code, "name": current_name or "",
                    "date": d, "tip_rev": 0.0, "disposal": 0.0,
                })
                t["tons"] += net
                if rate and not t["rate"]:
                    t["rate"] = rate
                if rate:
                    t["tip_rev"] += net * rate
                if not t["date"]:
                    t["date"] = d
                if not t["code"]:
                    t["code"] = effective_code
                if not t["name"]:
                    t["name"] = current_name or ""

        # Rolloff service lines → mark ticket as a rolloff (not a walk-in)
        for kw in ROLLOFF_SVC_KWS:
            if kw in flat_upper:
                ticket = data_vals[1] or "?"
                d = None
                for v in data_vals:
                    if v and re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", v):
                        d = v
                        break
                t = tickets.setdefault(ticket, {
                    "tons": 0.0, "rate": 0.0, "service": None,
                    "code": effective_code, "name": current_name or "",
                    "date": d, "tip_rev": 0.0, "disposal": 0.0,
                })
                t["service"] = kw
                if not t["date"]:
                    t["date"] = d
                if not t["name"]:
                    t["name"] = current_name or ""
                break

        # Disposal fees (mattress, appliance, freon, monitor, tire)
        for kw, price in DISPOSAL_FEES.items():
            if kw in flat_upper:
                ki = next((i for i, v in enumerate(data_vals)
                           if v and kw in v.upper()), -1)
                if ki != -1 and ki + 1 < len(data_vals):
                    try:
                        qty = float(data_vals[ki + 1])
                    except (TypeError, ValueError):
                        qty = 0.0
                    if qty > 0:
                        ticket = data_vals[1] or "?"
                        d = None
                        for v in data_vals:
                            if v and re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", v):
                                d = v
                                break
                        t = tickets.setdefault(ticket, {
                            "tons": 0.0, "rate": 0.0, "service": None,
                            "code": effective_code, "name": current_name or "",
                            "date": d, "tip_rev": 0.0, "disposal": 0.0,
                        })
                        t["disposal"] += qty * price
                        if not t["date"]:
                            t["date"] = d
                break

    return tickets


def is_walkin(t):
    code = t.get("code") or ""
    if code in IC_CODES:
        return False
    if code.startswith("ZZ"):
        return True  # ZZTNT and other cash walk-in codes
    return t.get("service") is None  # non-ZZ contractor drive-through


def main():
    print("Reading scale export…")
    with open(XLS_PATH, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()
    print(f"  read {len(text):,} chars")

    print("Parsing rows…")
    rows = parse_rows(text)
    print(f"  {len(rows):,} rows")

    print("Extracting tickets…")
    tickets = extract_tickets(rows)
    print(f"  {len(tickets):,} unique tickets parsed")

    # Filter to walk-ins in 2025 (Jan 1 2025 – Dec 31 2025 inclusive)
    walkins = []
    for tn, t in tickets.items():
        if not is_walkin(t):
            continue
        if not t["date"]:
            continue
        iso = mdy_to_iso(t["date"])
        if not iso or not ("2025-01-01" <= iso <= "2025-12-31"):
            continue
        walkins.append({
            "ticket": tn,
            "date_iso": iso,
            "code": t["code"],
            "name": t["name"],
            "tons": round(t["tons"], 3),
            "rate": round(t["rate"], 2),
            "tip_rev": round(t["tip_rev"], 2),
            "disposal": round(t["disposal"], 2),
            "total_rev": round(t["tip_rev"] + t["disposal"], 2),
        })

    walkins.sort(key=lambda r: (r["date_iso"], r["ticket"]))
    print(f"  {len(walkins):,} walk-in tickets in 2025")

    # Aggregates for headline numbers (sanity-check vs HELM)
    total_tons = sum(w["tons"] for w in walkins)
    total_tipRev = sum(w["tip_rev"] for w in walkins)
    total_disposal = sum(w["disposal"] for w in walkins)
    print(f"  2025 walk-in tons: {total_tons:,.2f}")
    print(f"  2025 walk-in tip revenue:  ${total_tipRev:,.2f}")
    print(f"  2025 walk-in disposal rev: ${total_disposal:,.2f}")
    print(f"  2025 walk-in total revenue: ${total_tipRev + total_disposal:,.2f}")

    write_workbook(walkins)
    print(f"\nWrote {OUT_PATH}")


# ─────────────────────────── Excel output ───────────────────────────

NAVY = "0e2a4f"
LIGHT = "f3f6fa"
GREEN = "1d8348"
RED = "b03a2e"


def thin_border():
    s = Side(style="thin", color="cccccc")
    return Border(left=s, right=s, top=s, bottom=s)


def write_workbook(walkins):
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    ws_tix = wb.create_sheet("Tickets")
    ws_mon = wb.create_sheet("Monthly")

    write_tickets_sheet(ws_tix, walkins)
    write_dashboard(ws_dash, walkins, ws_tix.max_row)
    write_monthly(ws_mon, walkins)

    wb.save(OUT_PATH)


def write_tickets_sheet(ws, walkins):
    headers = [
        "Date", "Ticket #", "Customer", "Code", "Tons",
        "Rate $/ton", "Tip Rev (actual)", "Disposal Rev", "Total Rev (actual)",
        "Adjusted Tons", "Adjusted Tip Rev", "Adjusted Total Rev", "Uplift $",
        "Under Min?",
    ]
    ws.append(headers)

    # Header style
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="ffffff", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    money_fmt = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    tons_fmt = "#,##0.00"
    int_fmt = "#,##0"

    for i, w in enumerate(walkins, start=2):
        ws.cell(row=i, column=1, value=w["date_iso"])
        ws.cell(row=i, column=2, value=w["ticket"])
        ws.cell(row=i, column=3, value=w["name"])
        ws.cell(row=i, column=4, value=w["code"])
        ws.cell(row=i, column=5, value=w["tons"]).number_format = tons_fmt
        ws.cell(row=i, column=6, value=w["rate"]).number_format = money_fmt
        ws.cell(row=i, column=7, value=w["tip_rev"]).number_format = money_fmt
        ws.cell(row=i, column=8, value=w["disposal"]).number_format = money_fmt
        ws.cell(row=i, column=9, value=w["total_rev"]).number_format = money_fmt

        # Formulas reference Dashboard!$B$3 (min tons) and Dashboard!$B$4 (rate override).
        ws.cell(row=i, column=10, value=f"=MAX(E{i},Dashboard!$B$3)").number_format = tons_fmt
        ws.cell(row=i, column=11, value=f"=J{i}*IF(Dashboard!$B$4>0,Dashboard!$B$4,F{i})").number_format = money_fmt
        ws.cell(row=i, column=12, value=f"=K{i}+H{i}").number_format = money_fmt
        ws.cell(row=i, column=13, value=f"=L{i}-I{i}").number_format = money_fmt
        ws.cell(row=i, column=14, value=f'=IF(E{i}<Dashboard!$B$3,"YES","")')

    widths = [12, 10, 32, 12, 9, 11, 14, 12, 14, 12, 14, 14, 12, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"


def write_dashboard(ws, walkins, ticket_rows):
    # Title
    ws["A1"] = "Walk-In Minimum Tonnage Analysis — 2025"
    ws["A1"].font = Font(size=16, bold=True, color="ffffff")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    # ── Controls (rows 2-4) — formulas in Tickets sheet point to B3 and B4 ──
    ws["A2"] = "CONTROLS"
    ws["A2"].font = Font(bold=True, color="ffffff", size=11)
    ws["A2"].fill = PatternFill("solid", fgColor="1f4e79")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    ws["A3"] = "Minimum Tonnage (tons)"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = 0.50
    ws["B3"].font = Font(bold=True, size=14, color="ffffff")
    ws["B3"].fill = PatternFill("solid", fgColor=GREEN)
    ws["B3"].alignment = Alignment(horizontal="center")
    ws["B3"].number_format = "0.00"
    ws["C3"] = "← MAIN KNOB — light tickets get billed at this weight × their rate"

    ws["A4"] = "Rate Override ($/ton, optional)"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = 0
    ws["B4"].font = Font(bold=True, size=14, color="ffffff")
    ws["B4"].fill = PatternFill("solid", fgColor="2874a6")
    ws["B4"].alignment = Alignment(horizontal="center")
    ws["B4"].number_format = '"$"#,##0.00'
    ws["C4"] = "← leave 0 to use each ticket's own rate; set a value to force that $/ton"

    for cell in ["A3", "B3", "C3", "A4", "B4", "C4"]:
        ws[cell].border = thin_border()

    # ── Headline KPIs (rows 6-15) ──
    last_tix_row = ticket_rows  # actual data ends at this row in Tickets sheet
    tix_range = f"Tickets!E2:E{last_tix_row}"
    actual_rev_range = f"Tickets!I2:I{last_tix_row}"
    adj_rev_range = f"Tickets!L2:L{last_tix_row}"
    uplift_range = f"Tickets!M2:M{last_tix_row}"
    under_range = f"Tickets!N2:N{last_tix_row}"

    ws["A6"] = "▸ HEADLINE IMPACT"
    ws["A6"].font = Font(bold=True, color="ffffff", size=11)
    ws["A6"].fill = PatternFill("solid", fgColor="1f4e79")
    ws.merge_cells("A6:F6")
    ws.row_dimensions[6].height = 22

    rows_kpi = [
        ("Total walk-in tickets (2025)",       f"=COUNTA(Tickets!B2:B{last_tix_row})", "#,##0"),
        ("Tickets under min threshold",        f'=COUNTIF({under_range},"YES")',        "#,##0"),
        ("% of tickets affected",              f'=COUNTIF({under_range},"YES")/COUNTA(Tickets!B2:B{last_tix_row})', "0.0%"),
        ("Total walk-in tons (actual)",        f"=SUM({tix_range})",                    "#,##0.00"),
        ("Current revenue (actual)",           f"=SUM({actual_rev_range})",             '"$"#,##0.00'),
        ("Adjusted revenue (with min)",        f"=SUM({adj_rev_range})",                '"$"#,##0.00'),
        ("Revenue uplift ($)",                 f"=SUM({uplift_range})",                 '"$"#,##0.00'),
        ("Revenue uplift (%)",                 f"=SUM({uplift_range})/SUM({actual_rev_range})", "0.0%"),
        ("Avg uplift per affected ticket",     f'=SUM({uplift_range})/MAX(1,COUNTIF({under_range},"YES"))', '"$"#,##0.00'),
    ]
    for i, (label, formula, fmt) in enumerate(rows_kpi, start=7):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=formula)
        c.number_format = fmt
        c.font = Font(bold=True, size=12, color=NAVY)
        c.alignment = Alignment(horizontal="right")
        for col in (1, 2):
            ws.cell(row=i, column=col).border = thin_border()

    # ── Distribution by tons bucket ──
    bucket_start = 17
    ws.cell(row=bucket_start, column=1, value="▸ DISTRIBUTION — walk-in tickets by load weight").font = Font(bold=True, color="ffffff", size=11)
    ws.cell(row=bucket_start, column=1).fill = PatternFill("solid", fgColor="1f4e79")
    ws.merge_cells(f"A{bucket_start}:F{bucket_start}")
    ws.row_dimensions[bucket_start].height = 22

    headers = ["Load weight bucket", "# Tickets", "% of Tickets", "Tons", "Current Rev", "Uplift @ current min"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=bucket_start + 1, column=i, value=h)
        c.font = Font(bold=True, color="ffffff")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    buckets = [
        ("0.00 – 0.10",       0.00, 0.10),
        ("0.10 – 0.25",       0.10, 0.25),
        ("0.25 – 0.50",       0.25, 0.50),
        ("0.50 – 1.00",       0.50, 1.00),
        ("1.00 – 2.00",       1.00, 2.00),
        ("2.00 – 5.00",       2.00, 5.00),
        ("5.00+",             5.00, 9999.0),
    ]
    for i, (label, lo, hi) in enumerate(buckets):
        r = bucket_start + 2 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2,
                value=f'=COUNTIFS(Tickets!E2:E{last_tix_row},">="&{lo},Tickets!E2:E{last_tix_row},"<"&{hi})'
        ).number_format = "#,##0"
        ws.cell(row=r, column=3,
                value=f'=B{r}/COUNTA(Tickets!B2:B{last_tix_row})'
        ).number_format = "0.0%"
        ws.cell(row=r, column=4,
                value=f'=SUMIFS(Tickets!E2:E{last_tix_row},Tickets!E2:E{last_tix_row},">="&{lo},Tickets!E2:E{last_tix_row},"<"&{hi})'
        ).number_format = "#,##0.00"
        ws.cell(row=r, column=5,
                value=f'=SUMIFS(Tickets!I2:I{last_tix_row},Tickets!E2:E{last_tix_row},">="&{lo},Tickets!E2:E{last_tix_row},"<"&{hi})'
        ).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6,
                value=f'=SUMIFS(Tickets!M2:M{last_tix_row},Tickets!E2:E{last_tix_row},">="&{lo},Tickets!E2:E{last_tix_row},"<"&{hi})'
        ).number_format = '"$"#,##0.00'
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = thin_border()

    # Totals row
    total_r = bucket_start + 2 + len(buckets)
    ws.cell(row=total_r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_r, column=2, value=f"=SUM(B{bucket_start+2}:B{total_r-1})").number_format = "#,##0"
    ws.cell(row=total_r, column=3, value=f"=SUM(C{bucket_start+2}:C{total_r-1})").number_format = "0.0%"
    ws.cell(row=total_r, column=4, value=f"=SUM(D{bucket_start+2}:D{total_r-1})").number_format = "#,##0.00"
    ws.cell(row=total_r, column=5, value=f"=SUM(E{bucket_start+2}:E{total_r-1})").number_format = '"$"#,##0.00'
    ws.cell(row=total_r, column=6, value=f"=SUM(F{bucket_start+2}:F{total_r-1})").number_format = '"$"#,##0.00'
    for col in range(1, 7):
        c = ws.cell(row=total_r, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.border = thin_border()

    # ── Quick-comparison table — what would 4 different min thresholds yield? ──
    cmp_start = total_r + 3
    ws.cell(row=cmp_start, column=1, value="▸ QUICK COMPARE — uplift @ different min thresholds (your ticket rates)").font = Font(bold=True, color="ffffff", size=11)
    ws.cell(row=cmp_start, column=1).fill = PatternFill("solid", fgColor="1f4e79")
    ws.merge_cells(f"A{cmp_start}:F{cmp_start}")
    ws.row_dimensions[cmp_start].height = 22

    headers2 = ["Min Tons", "Tickets affected", "Tons added (billable)", "Uplift $", "Uplift %", ""]
    for i, h in enumerate(headers2, start=1):
        c = ws.cell(row=cmp_start + 1, column=i, value=h)
        c.font = Font(bold=True, color="ffffff")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    # These four scenarios are computed statically and embedded as values
    # (Excel can't easily run an array of MIN/MAX calcs across mass-of-rows
    # with a different parameter per row without LET/dynamic-array.)
    static = _scenarios(walkins, [0.25, 0.50, 0.75, 1.00])
    for i, sc in enumerate(static):
        r = cmp_start + 2 + i
        ws.cell(row=r, column=1, value=sc["min_tons"]).number_format = "0.00"
        ws.cell(row=r, column=2, value=sc["affected"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(sc["tons_added"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=round(sc["uplift"], 2)).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5, value=sc["uplift_pct"]).number_format = "0.0%"
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = thin_border()

    # Column widths
    widths = {"A": 36, "B": 18, "C": 55, "D": 18, "E": 16, "F": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _scenarios(walkins, thresholds):
    """Compute static comparison rows so the dashboard has at-a-glance context."""
    out = []
    base_rev = sum(w["total_rev"] for w in walkins)
    for thr in thresholds:
        affected = 0
        tons_added = 0.0
        uplift = 0.0
        for w in walkins:
            if w["tons"] < thr:
                affected += 1
                tons_added += (thr - w["tons"])
                uplift += (thr - w["tons"]) * w["rate"]
        out.append({
            "min_tons": thr,
            "affected": affected,
            "tons_added": tons_added,
            "uplift": uplift,
            "uplift_pct": (uplift / base_rev) if base_rev else 0,
        })
    return out


def write_monthly(ws, walkins):
    by_month = defaultdict(lambda: {"tix": 0, "tons": 0.0, "rev": 0.0, "under_05": 0, "under_05_rev": 0.0})
    for w in walkins:
        ym = w["date_iso"][:7]
        b = by_month[ym]
        b["tix"] += 1
        b["tons"] += w["tons"]
        b["rev"] += w["total_rev"]
        if w["tons"] < 0.5:
            b["under_05"] += 1
            b["under_05_rev"] += w["total_rev"]

    headers = ["Month", "Walk-in tickets", "Walk-in tons", "Walk-in revenue",
               "Tickets <0.5t", "% tickets <0.5t", "Rev from <0.5t",
               "Modeled uplift @ 0.5t min"]
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="ffffff")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    money_fmt = '"$"#,##0.00'
    for i, ym in enumerate(sorted(by_month.keys()), start=2):
        b = by_month[ym]
        # uplift for this month at 0.5t min
        uplift = 0.0
        for w in walkins:
            if w["date_iso"][:7] != ym:
                continue
            if w["tons"] < 0.5:
                uplift += (0.5 - w["tons"]) * w["rate"]

        ws.cell(row=i, column=1, value=ym)
        ws.cell(row=i, column=2, value=b["tix"]).number_format = "#,##0"
        ws.cell(row=i, column=3, value=round(b["tons"], 2)).number_format = "#,##0.00"
        ws.cell(row=i, column=4, value=round(b["rev"], 2)).number_format = money_fmt
        ws.cell(row=i, column=5, value=b["under_05"]).number_format = "#,##0"
        ws.cell(row=i, column=6, value=b["under_05"] / b["tix"] if b["tix"] else 0).number_format = "0.0%"
        ws.cell(row=i, column=7, value=round(b["under_05_rev"], 2)).number_format = money_fmt
        ws.cell(row=i, column=8, value=round(uplift, 2)).number_format = money_fmt
        for col in range(1, 9):
            ws.cell(row=i, column=col).border = thin_border()

    # Totals row
    last_row = ws.max_row
    tot_r = last_row + 1
    ws.cell(row=tot_r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tot_r, column=2, value=f"=SUM(B2:B{last_row})").number_format = "#,##0"
    ws.cell(row=tot_r, column=3, value=f"=SUM(C2:C{last_row})").number_format = "#,##0.00"
    ws.cell(row=tot_r, column=4, value=f"=SUM(D2:D{last_row})").number_format = money_fmt
    ws.cell(row=tot_r, column=5, value=f"=SUM(E2:E{last_row})").number_format = "#,##0"
    ws.cell(row=tot_r, column=6, value=f"=E{tot_r}/B{tot_r}").number_format = "0.0%"
    ws.cell(row=tot_r, column=7, value=f"=SUM(G2:G{last_row})").number_format = money_fmt
    ws.cell(row=tot_r, column=8, value=f"=SUM(H2:H{last_row})").number_format = money_fmt
    for col in range(1, 9):
        c = ws.cell(row=tot_r, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.border = thin_border()

    widths = [10, 16, 14, 18, 14, 16, 18, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    main()
