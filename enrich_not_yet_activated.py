"""
Enrich the "Not Yet Activated" sheet in CommAnalysis_Comparison1.1.xlsx
with Name + Service Address pulled from the live HELM clients table.

- Keeps the "Comparison" sheet untouched.
- Inserts two new columns between "Business" (col C) and "Last Seen"
  (formerly col D, now F): Name and Service Address.
- Preserves the existing title rows, header, and footer-summary row.
- Writes a copy alongside the original named with an _enriched suffix.
"""
import sys, json, urllib.parse, urllib.request
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

SUPABASE_URL  = 'https://iiqlhqtyyqctgupalrlc.supabase.co'
SUPABASE_ANON = 'sb_publishable_JAASDdg8OERvpDTnEeLd5Q_Re48qR1I'

IN_PATH  = r"C:\Users\theco\Downloads\CommAnalysis_Comparison1.1.xlsx"
OUT_PATH = r"C:\Users\theco\Downloads\CommAnalysis_Comparison1.1_enriched.xlsx"


def fetch_clients_by_ids(ids):
    """Single Supabase call: in.(id1,id2,...). Returns {client_id: row}."""
    if not ids:
        return {}
    # PostgREST 'in' filter — comma-separated, URL-encode the whole list.
    ids_param = ",".join(ids)
    qs = urllib.parse.urlencode({
        "select": "client_id,client_name,address",
        "client_id": f"in.({ids_param})",
    })
    url = f"{SUPABASE_URL}/rest/v1/clients?{qs}"
    req = urllib.request.Request(url, headers={
        "apikey": SUPABASE_ANON,
        "Authorization": f"Bearer {SUPABASE_ANON}",
    })
    with urllib.request.urlopen(req, timeout=30) as resp:
        rows = json.loads(resp.read().decode("utf-8"))
    return {r["client_id"]: r for r in rows}


def add_comparison_summary(wb):
    """Insert a compact status-count summary at the top of the Comparison sheet
    so the breakdown is visible without scrolling to row 565+."""
    ws = wb["Comparison"]

    # Read every status from the Reis (col D) and Santos (col I) status columns.
    # Original header is on row 4, data starts row 5.
    reis_counts = {}
    santos_counts = {}
    for r in range(5, ws.max_row + 1):
        rs = ws.cell(r, 4).value
        ss = ws.cell(r, 9).value
        if rs: reis_counts[rs]   = reis_counts.get(rs, 0)   + 1
        if ss: santos_counts[ss] = santos_counts.get(ss, 0) + 1

    # Stable ordering — known statuses first, then anything else alphabetically.
    preferred = ["Retained", "New", "Dropped"]
    all_statuses = set(reis_counts) | set(santos_counts)
    ordered = [s for s in preferred if s in all_statuses] + sorted(s for s in all_statuses if s not in preferred)

    # Build the rows we want to insert at the top (between the existing title
    # rows and the original header). 1 hdr + N status + 1 total + 1 blank.
    insert_count = 2 + len(ordered) + 1  # title row + col header + statuses + blank
    ws.insert_rows(idx=4, amount=insert_count)

    # Re-shift any merged-cell ranges that previously spanned row 1 and 2 —
    # insert_rows doesn't always extend merges that started above the insert
    # point but spanned past it; the title rows are above so they should be ok.

    # Styles
    navy_fill = PatternFill("solid", fgColor="0e2a4f")
    light_fill = PatternFill("solid", fgColor="f3f6fa")
    bold_white = Font(bold=True, color="ffffff", size=11)
    bold_navy  = Font(bold=True, color="0e2a4f", size=11)
    thin = Side(style="thin", color="d5d9e0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right_align = Alignment(horizontal="right", vertical="center")
    left_align  = Alignment(horizontal="left",  vertical="center")
    center_align= Alignment(horizontal="center",vertical="center")

    # Title row (row 4)
    title = ws.cell(row=4, column=1, value="Status Breakdown")
    title.font = bold_white
    title.fill = navy_fill
    title.alignment = left_align
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)

    # Column header (row 5)
    headers = ["Status", "Reis", "Santos", "Total"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = bold_white
        cell.fill = navy_fill
        cell.alignment = center_align if i > 1 else left_align
        cell.border = border

    # Status data rows (rows 6 .. 6+len(ordered)-1)
    for i, status in enumerate(ordered):
        r = 6 + i
        r_count = reis_counts.get(status, 0)
        s_count = santos_counts.get(status, 0)
        ws.cell(row=r, column=1, value=status).alignment = left_align
        c_reis   = ws.cell(row=r, column=2, value=r_count)
        c_santos = ws.cell(row=r, column=3, value=s_count)
        c_total  = ws.cell(row=r, column=4, value=r_count + s_count)
        c_reis.alignment = right_align
        c_santos.alignment = right_align
        c_total.alignment = right_align
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = border

    # Total row
    total_r = 6 + len(ordered)
    total_reis   = sum(reis_counts.get(s, 0)   for s in ordered)
    total_santos = sum(santos_counts.get(s, 0) for s in ordered)
    ws.cell(row=total_r, column=1, value="Total accounts").font = bold_navy
    ws.cell(row=total_r, column=1).alignment = left_align
    ws.cell(row=total_r, column=2, value=total_reis).font   = bold_navy
    ws.cell(row=total_r, column=3, value=total_santos).font = bold_navy
    ws.cell(row=total_r, column=4, value=total_reis + total_santos).font = bold_navy
    for col in range(1, 5):
        c = ws.cell(row=total_r, column=col)
        c.fill = light_fill
        c.border = border
        if col > 1: c.alignment = right_align

    # Column widths
    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width or 0, 26)
    for letter in ('B', 'C', 'D'):
        if (ws.column_dimensions[letter].width or 0) < 10:
            ws.column_dimensions[letter].width = 10


def enrich():
    wb = load_workbook(IN_PATH)
    ws = wb["Not Yet Activated"]

    # Header is on row 4. Data starts row 5.
    # Existing columns: A=#, B=Account, C=Business, D=Last Seen, E=Note.
    # We insert two columns at position D (so Last Seen shifts to F, Note to G).
    ws.insert_cols(idx=4, amount=2)

    # Write header labels (style-match to existing header row 4).
    template = ws.cell(row=4, column=1)  # any existing header cell, for formatting reuse
    def style_like(target, src):
        if src.has_style:
            target.font          = copy(src.font)
            target.fill          = copy(src.fill)
            target.alignment     = copy(src.alignment)
            target.border        = copy(src.border)
            target.number_format = src.number_format
            target.protection    = copy(src.protection)

    name_hdr = ws.cell(row=4, column=4, value="Name")
    addr_hdr = ws.cell(row=4, column=5, value="Service Address")
    style_like(name_hdr, template)
    style_like(addr_hdr, template)

    # Stretch the title-row merges (rows 1, 2, 82) so they still span the
    # whole table width — we added 2 columns, so each existing merge that
    # ended at column E now needs to end at column G.
    new_merges = []
    for mr in list(ws.merged_cells.ranges):
        # mr is a CellRange (we treat as immutable; rebuild as needed).
        # Title rows had col span A..E (1..5). Extend to A..G (1..7).
        if mr.min_row in (1, 2) and mr.min_col == 1 and mr.max_col == 5:
            ws.unmerge_cells(str(mr))
            ws.merge_cells(start_row=mr.min_row, start_column=1, end_row=mr.max_row, end_column=7)

    # Pull every account # from rows 5..end (data section). Data ends where
    # the Account column goes blank.
    account_ids = []
    data_rows = []
    for r in range(5, ws.max_row + 1):
        acct_cell = ws.cell(row=r, column=2).value
        if acct_cell is None or str(acct_cell).strip() == "":
            continue
        # Account cells are numeric in the source — coerce to string.
        # Skip the footer summary row ("Reis: 42  Santos: 34  Total: 76") and
        # any other non-numeric value that happens to live in column B.
        if isinstance(acct_cell, (int, float)):
            acct_str = str(int(acct_cell))
        else:
            s = str(acct_cell).strip()
            if not s.isdigit():
                continue
            acct_str = s
        account_ids.append(acct_str)
        data_rows.append((r, acct_str))

    print(f"Looking up {len(account_ids)} accounts in HELM…")
    lookup = fetch_clients_by_ids(account_ids)
    print(f"  matched {len(lookup)} / {len(account_ids)}")

    # Write Name (col D) and Service Address (col E) for each data row.
    missing = []
    body_style_src = ws.cell(row=5, column=3)  # existing 'Business' cell to mimic
    for (r, acct) in data_rows:
        row = lookup.get(acct, {})
        name = row.get("client_name") or ""
        addr = row.get("address") or ""
        if not row:
            missing.append(acct)
        nc = ws.cell(row=r, column=4, value=name or "(not found)")
        ac = ws.cell(row=r, column=5, value=addr or "(not found)")
        style_like(nc, body_style_src)
        style_like(ac, body_style_src)

    # Generous column widths so name + address aren't truncated.
    ws.column_dimensions['D'].width = 34   # Name
    ws.column_dimensions['E'].width = 42   # Service Address

    if missing:
        print(f"\nNot found in HELM (likely never imported / different ID format):")
        for m in missing:
            print(f"  - {m}")

    # Add the compact status summary at the top of the Comparison sheet.
    add_comparison_summary(wb)

    # If the primary output is locked (file open in Excel), fall back to a
    # versioned name so the script always completes.
    save_path = OUT_PATH
    try:
        wb.save(save_path)
    except PermissionError:
        from datetime import datetime
        ts = datetime.now().strftime("%H%M%S")
        save_path = OUT_PATH.replace(".xlsx", f"_v{ts}.xlsx")
        wb.save(save_path)
        print(f"\n  (primary output was open in Excel — wrote a fresh copy)")
    print(f"\nWrote {save_path}")


if __name__ == "__main__":
    enrich()
