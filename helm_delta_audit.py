"""
helm_delta_audit.py — cross-reference HELM clients against the Delta PDF exports.

Buckets every HELM client into one of three states:
  1. In delta WITH route notes      (covered by 'delta export retry 2.pdf')
  2. In delta WITHOUT route notes   (covered by 'delta export retry.pdf' only)
  3. Totally missing from delta     (not in either PDF)

Outputs:
  audit_in_delta_with_routes.csv
  audit_in_delta_no_routes.csv
  audit_helm_only.csv
  audit_summary.txt

Acct numbers starting with 1 (1xxxxx) are filtered out per project convention.
"""
import pdfplumber, re, json, csv, collections, sys
from urllib.request import Request, urlopen
from urllib.error import HTTPError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- config ------------------------------------------------------------
ROOT = r"C:\Users\theco\OneDrive\Desktop\Nantucket\helm-app"
RETRY_PDF  = ROOT + r"\delta export retry.pdf"      # all accts, no routes (5,588)
RETRY2_PDF = ROOT + r"\delta export retry 2.pdf"    # subset with routes (1,564)

SUPABASE_URL  = 'https://iiqlhqtyyqctgupalrlc.supabase.co'
SUPABASE_ANON = 'sb_publishable_JAASDdg8OERvpDTnEeLd5Q_Re48qR1I'


# --- PDF acct extraction -----------------------------------------------
def extract_accts_retry(pdf_path):
    """retry.pdf — compact 'All Accounts' format.
    Header line per client: ACCT[*] CODE NAME, ADDR / CITY DATE [DAYS] TOTAL
    """
    with pdfplumber.open(pdf_path) as pdf:
        full = '\n'.join((p.extract_text() or '') for p in pdf.pages)
    accts = set()
    # 6 digits, optional *, optional space, code (alphanumeric 2-5), then a space+anything
    for m in re.finditer(r'^(\d{6})\*?\s*[A-Za-z0-9]{2,5}\s+\S', full, re.M):
        accts.add(m.group(1))
    return accts


def extract_accts_retry2(pdf_path):
    """retry 2.pdf — 'Master List All Accounts' format with dotted block separators.
    Returns:
      all_accts        — every acct in the PDF (regardless of whether it has route records)
      accts_with_routes — subset that has at least one 'N Route Records X-NN POS' line
    """
    with pdfplumber.open(pdf_path) as pdf:
        full = '\n'.join((p.extract_text() or '') for p in pdf.pages)

    # Acct line in this format. Status codes can be:
    #   uppercase (RAMA, REMA, XX) — residential / system
    #   mixed-case (C1Ar, C2Ar, C1Ad) — commercial
    # So allow alphanumerics + mixed case in the {2,5} group.
    all_accts = set()
    for m in re.finditer(r'^(\d{6})\s+[A-Za-z0-9]{2,5}\b', full, re.M):
        all_accts.add(m.group(1))

    # Now split by the dotted separator and check each block for route records
    sep_re = re.compile(r'^\s*(?:\.\s+){10,}\.?\s*$', re.M)
    blocks = sep_re.split(full)
    accts_with_routes = set()
    for blk in blocks:
        m = re.search(r'^(\d{6})\s+[A-Za-z0-9]{2,5}\b', blk, re.M)
        if not m:
            continue
        acct = m.group(1)
        # has at least one "N Route Records X-NN POS" header?
        if re.search(r'\d+ Route Records?\s+[A-Z]-\d+\s+\d+', blk):
            accts_with_routes.add(acct)
    return all_accts, accts_with_routes


# --- HELM fetch --------------------------------------------------------
def fetch_helm():
    headers = {'apikey': SUPABASE_ANON, 'Authorization': f'Bearer {SUPABASE_ANON}'}
    rows = []
    offset = 0
    while True:
        url = (f'{SUPABASE_URL}/rest/v1/clients'
               f'?select=client_id,client_name,address,phone,status,service_day,route'
               f'&order=client_id&offset={offset}&limit=1000')
        req = Request(url, headers=headers)
        try:
            with urlopen(req, timeout=30) as resp:
                batch = json.loads(resp.read().decode('utf-8'))
        except HTTPError as e:
            print(f'Supabase fetch failed (HTTP {e.code}): {e.read().decode("utf-8", errors="replace")}')
            sys.exit(1)
        if not batch:
            break
        rows.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    return rows


# --- audit -------------------------------------------------------------
def bucket_prefix(acct_id):
    if not acct_id or not acct_id.isdigit():
        return 'non-numeric / legacy'
    n = int(acct_id)
    if   100000 <= n < 200000: return '1xx Reis system (filtered)'
    elif 200000 <= n < 210000: return '20xxxx Reis (residential range)'
    elif 210000 <= n < 300000: return '21-29xxxx Reis other'
    elif 300000 <= n < 400000: return '3xxxxx SANTOS'
    elif 400000 <= n < 500000: return '4xxxxx other'
    elif n < 100000:           return '<100000 short legacy'
    return '>=500000 other'


def to_row(h):
    return {
        'acct': h['client_id'],
        'name': h.get('client_name') or '',
        'address': h.get('address') or '',
        'status': h.get('status') or '',
        'service_day': h.get('service_day') or '',
        'route': '' if h.get('route') is None else h.get('route'),
        'phone': h.get('phone') or '',
        'prefix_bucket': bucket_prefix(h['client_id']),
    }


def write_csv(path, rows):
    cols = ['acct', 'name', 'address', 'status', 'service_day', 'route', 'phone', 'prefix_bucket']
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main():
    print('Extracting acct numbers from retry.pdf (5588 expected, no route records)...')
    retry_accts = extract_accts_retry(RETRY_PDF)
    print(f'  {len(retry_accts)} accts')

    print('Extracting acct numbers from retry 2.pdf (1564 expected, with routes)...')
    retry2_accts, retry2_with_routes = extract_accts_retry2(RETRY2_PDF)
    print(f'  {len(retry2_accts)} accts, {len(retry2_with_routes)} with route records')

    print('Fetching HELM clients via Supabase...')
    helm = fetch_helm()
    print(f'  {len(helm)} clients')

    in_delta_with_routes = []
    in_delta_no_routes   = []
    not_in_delta         = []
    skipped_1xxxxx       = 0

    delta_union = retry_accts | retry2_accts

    for h in helm:
        acct = h['client_id']
        if acct and acct.isdigit() and acct.startswith('1') and len(acct) == 6:
            skipped_1xxxxx += 1
            continue
        if acct in retry2_with_routes:
            in_delta_with_routes.append(to_row(h))
        elif acct in delta_union:
            in_delta_no_routes.append(to_row(h))
        else:
            not_in_delta.append(to_row(h))

    # Sort each bucket by acct
    in_delta_with_routes.sort(key=lambda r: r['acct'])
    in_delta_no_routes.sort(key=lambda r: r['acct'])
    not_in_delta.sort(key=lambda r: r['acct'])

    write_csv('audit_in_delta_with_routes.csv', in_delta_with_routes)
    write_csv('audit_in_delta_no_routes.csv',   in_delta_no_routes)
    write_csv('audit_helm_only.csv',            not_in_delta)

    # Reverse direction: delta accts not in HELM
    helm_acct_set = {h['client_id'] for h in helm}
    delta_only = sorted(a for a in delta_union if a not in helm_acct_set and not a.startswith('1'))

    with open('audit_delta_only.csv', 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['acct', 'in_retry', 'in_retry2', 'in_retry2_with_routes'])
        for a in delta_only:
            w.writerow([
                a,
                'yes' if a in retry_accts else 'no',
                'yes' if a in retry2_accts else 'no',
                'yes' if a in retry2_with_routes else 'no',
            ])

    # Bucket summaries
    def buckets(rows):
        c = collections.Counter(r['prefix_bucket'] for r in rows)
        return c

    not_in_delta_b = buckets(not_in_delta)
    in_delta_no_r_b = buckets(in_delta_no_routes)
    in_delta_with_r_b = buckets(in_delta_with_routes)

    # Status breakdown for the actionable buckets (Active vs Paused)
    def status_breakdown(rows):
        return collections.Counter(r['status'] or '?' for r in rows)

    # Summary
    lines = [
        'HELM <-> Delta cross-reference audit',
        f'Generated: {__import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        '',
        'Sources:',
        f'  HELM clients (Supabase):     {len(helm)}',
        f'  Delta retry (all accts):     {len(retry_accts)} accts',
        f'  Delta retry 2 (master+rts):  {len(retry2_accts)} accts ({len(retry2_with_routes)} with route records)',
        f'  Union of both PDFs:          {len(delta_union)} unique accts',
        '',
        f'1xxxxx HELM accts skipped:     {skipped_1xxxxx}',
        '',
        '--- HELM bucketed ---------------------------------------------',
        f'In delta WITH route notes:           {len(in_delta_with_routes)}',
        f'In delta WITHOUT route notes:        {len(in_delta_no_routes)}',
        f'Totally MISSING from delta:          {len(not_in_delta)}',
        '',
        '--- "Totally missing" — by acct prefix ------------------------',
    ]
    for b, n in not_in_delta_b.most_common():
        lines.append(f'  {b:<35} {n}')

    lines.append('')
    lines.append('--- "Totally missing" — by HELM status ------------------------')
    sb_missing = status_breakdown(not_in_delta)
    for s, n in sb_missing.most_common():
        lines.append(f'  {s:<35} {n}')

    lines.append('')
    lines.append('--- "In delta, missing route notes" — by status ---------------')
    sb_no_routes = status_breakdown(in_delta_no_routes)
    for s, n in sb_no_routes.most_common():
        lines.append(f'  {s:<35} {n}')

    lines.append('')
    lines.append('--- "In delta with routes" — by status ------------------------')
    sb_with_routes = status_breakdown(in_delta_with_routes)
    for s, n in sb_with_routes.most_common():
        lines.append(f'  {s:<35} {n}')

    lines.append('')
    lines.append('--- reverse direction ----------------------------------------')
    lines.append(f'Delta accts NOT in HELM (would CREATE):  {len(delta_only)}')

    summary = '\n'.join(lines)
    with open('audit_summary.txt', 'w', encoding='utf-8') as f:
        f.write(summary)
    # ASCII-only stdout for Windows cp1252 console
    print(summary.encode('ascii', 'replace').decode('ascii'))
    print()

    # Build the bundled XLSX workbook
    wb = Workbook()
    NAVY = 'FF0F1F3D'
    LIGHT_GRAY = 'FFF7F8FA'

    def add_summary_sheet():
        ws = wb.active
        ws.title = 'Summary'
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 18
        ws['A1'] = 'HELM <-> Delta cross-reference audit'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor=NAVY)
        ws.merge_cells('A1:B1')

        row = 3
        rows_to_add = [
            ('Sources', ''),
            ('  HELM clients (Supabase)',         len(helm)),
            ('  Delta retry — all accts',         len(retry_accts)),
            ('  Delta retry 2 — total accts',     len(retry2_accts)),
            ('  Delta retry 2 — with route records', len(retry2_with_routes)),
            ('  Union of both PDFs',              len(delta_union)),
            ('', ''),
            ('1xxxxx HELM accts skipped',         skipped_1xxxxx),
            ('', ''),
            ('HELM bucketed', ''),
            ('  In delta WITH route notes',       len(in_delta_with_routes)),
            ('  In delta WITHOUT route notes',    len(in_delta_no_routes)),
            ('  Totally MISSING from delta',      len(not_in_delta)),
            ('', ''),
            ('Reverse direction', ''),
            ('  Delta accts NOT in HELM (would CREATE)', len(delta_only)),
        ]
        for label, val in rows_to_add:
            ws.cell(row=row, column=1, value=label)
            if isinstance(val, int):
                ws.cell(row=row, column=2, value=val)
                ws.cell(row=row, column=2).font = Font(bold=True)
            else:
                ws.cell(row=row, column=2, value=val)
            if label and not label.startswith('  ') and val == '':
                # Section header
                ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFFFF')
                ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=NAVY)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1

        # "Totally missing" prefix bucket sub-table
        row += 1
        ws.cell(row=row, column=1, value='"Totally missing" — by acct prefix bucket').font = Font(bold=True, color='FFFFFFFF')
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=NAVY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        for b, n in not_in_delta_b.most_common():
            ws.cell(row=row, column=1, value='  ' + b)
            ws.cell(row=row, column=2, value=n).font = Font(bold=True)
            row += 1

        # Status breakdowns
        for label, c in [
            ('"Totally missing" — by HELM status', sb_missing),
            ('"In delta no route notes" — by HELM status', sb_no_routes),
            ('"In delta with routes" — by HELM status', sb_with_routes),
        ]:
            row += 1
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, color='FFFFFFFF')
            ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=NAVY)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            for s, n in c.most_common():
                ws.cell(row=row, column=1, value='  ' + s)
                ws.cell(row=row, column=2, value=n).font = Font(bold=True)
                row += 1

    def add_detail_sheet(name, rows):
        ws = wb.create_sheet(name)
        cols = ['acct', 'name', 'address', 'status', 'service_day', 'route', 'phone', 'prefix_bucket']
        # Header row
        for c, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill('solid', fgColor=NAVY)
            cell.alignment = Alignment(horizontal='left')
        # Data rows
        for r, row in enumerate(rows, start=2):
            for c, col in enumerate(cols, start=1):
                ws.cell(row=r, column=c, value=row.get(col, ''))
            if r % 2 == 0:
                for c in range(1, len(cols) + 1):
                    ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=LIGHT_GRAY)
        # Auto-width approximation
        widths = [10, 32, 38, 10, 22, 8, 16, 30]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{max(2, len(rows) + 1)}'
        ws.cell(row=1, column=1).comment = None  # noqa

    def add_delta_only_sheet():
        ws = wb.create_sheet('Delta only (would CREATE)')
        cols = ['acct', 'in_retry', 'in_retry2', 'in_retry2_with_routes']
        for c, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill('solid', fgColor=NAVY)
        for r, a in enumerate(delta_only, start=2):
            ws.cell(row=r, column=1, value=a)
            ws.cell(row=r, column=2, value='yes' if a in retry_accts else 'no')
            ws.cell(row=r, column=3, value='yes' if a in retry2_accts else 'no')
            ws.cell(row=r, column=4, value='yes' if a in retry2_with_routes else 'no')
            if r % 2 == 0:
                for c in range(1, 5):
                    ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=LIGHT_GRAY)
        for i, w in enumerate([10, 12, 12, 22], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

    add_summary_sheet()
    add_detail_sheet('In delta with routes',     in_delta_with_routes)
    add_detail_sheet('In delta NO route notes',  in_delta_no_routes)
    add_detail_sheet('HELM only (missing)',      not_in_delta)
    add_delta_only_sheet()

    xlsx_path = 'HELM_Delta_Audit.xlsx'
    wb.save(xlsx_path)
    print(f'Bundled workbook -> {xlsx_path}')
    print()
    print('CSVs (also written): audit_in_delta_with_routes.csv, audit_in_delta_no_routes.csv,')
    print('                     audit_helm_only.csv, audit_delta_only.csv')


if __name__ == '__main__':
    main()
